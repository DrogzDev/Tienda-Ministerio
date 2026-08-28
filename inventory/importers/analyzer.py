from django.core.exceptions import ValidationError
from openpyxl import load_workbook

from inventory.selectors import (
    find_matching_products,
    get_product_by_code,
    resolve_category,
    resolve_unit,
)
from inventory.utils import normalize_header, normalize_text, parse_decimal


HEADER_ALIASES = {
    "CODIGO_PRODUCTO": {"CODIGO_PRODUCTO", "CODIGO", "COD_PRODUCTO"},
    "DESCRIPCION": {"DESCRIPCION", "PRODUCTO"},
    "CATEGORIA": {"CATEGORIA"},
    "UNIDAD_MEDIDA": {
        "UNIDAD_MEDIDA",
        "UND_DE_MEDIDA",
        "UNIDAD_DE_MEDIDA",
        "UND_MEDIDA",
    },
    "CANTIDAD": {"CANTIDAD"},
    "STOCK_MINIMO": {"STOCK_MINIMO", "MINIMO", "EXISTENCIA_MINIMA"},
    "OBSERVACIONES": {"OBSERVACIONES", "OBSERVACION", "NOTAS"},
}

REQUIRED_COLUMNS = {
    "DESCRIPCION",
    "CATEGORIA",
    "UNIDAD_MEDIDA",
    "CANTIDAD",
}


def _canonical_header(value):
    normalized = normalize_header(value)
    for canonical, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return canonical
    return normalized


def _sheet_size(sheet):
    """
    Devuelve dimensiones seguras incluso cuando openpyxl no trae
    max_row/max_column calculados en algunos archivos generados.
    """
    max_row = sheet.max_row or 0
    max_column = sheet.max_column or 0

    if max_row <= 0 or max_column <= 0:
        # Forzar cálculo de dimensiones en hojas con metadatos incompletos.
        dimension = sheet.calculate_dimension()
        if dimension:
            from openpyxl.utils.cell import range_boundaries
            min_col, min_row, max_col, max_row_calc = range_boundaries(dimension)
            max_row = max_row or max_row_calc
            max_column = max_column or max_col

    if max_row <= 0:
        max_row = 1
    if max_column <= 0:
        max_column = 1

    return int(max_row), int(max_column)


def _find_header_row(sheet):
    max_row, max_column = _sheet_size(sheet)

    for row_number in range(1, min(max_row, 10) + 1):
        values = [
            sheet.cell(row=row_number, column=col).value
            for col in range(1, max_column + 1)
        ]
        canonical = {
            _canonical_header(value)
            for value in values
            if value not in (None, "")
        }
        if REQUIRED_COLUMNS.issubset(canonical):
            return row_number

    raise ValidationError(
        "No se encontró una cabecera válida. La plantilla debe contener: "
        "DESCRIPCION, CATEGORIA, UNIDAD_MEDIDA y CANTIDAD."
    )


def _build_column_map(sheet, header_row):
    columns = {}
    _, max_column = _sheet_size(sheet)

    for col in range(1, max_column + 1):
        value = sheet.cell(row=header_row, column=col).value
        if value in (None, ""):
            continue
        canonical = _canonical_header(value)
        if canonical in HEADER_ALIASES:
            columns[canonical] = col

    missing = REQUIRED_COLUMNS - set(columns)
    if missing:
        raise ValidationError(
            f"Faltan columnas obligatorias: {', '.join(sorted(missing))}."
        )
    return columns


def _cell(sheet, row, columns, name):
    col = columns.get(name)
    return sheet.cell(row=row, column=col).value if col else None


def _decimal_to_string(number):
    if number is None:
        return None
    return format(number, "f")


def analyze_inventory_excel(*, file_or_path, section, sheet_name="Carga Inventario"):
    """
    Analiza el Excel SIN modificar la base de datos.

    Devuelve filas READY_EXISTING, READY_NEW, REVIEW o ERROR.
    """
    # read_only=False es intencional. Algunos XLSX generados por distintas
    # herramientas no exponen max_row/max_column correctamente en modo
    # read_only y openpyxl puede devolver None. Para inventarios de cientos
    # o pocos miles de filas el consumo de memoria es perfectamente razonable.
    if hasattr(file_or_path, "seek"):
        file_or_path.seek(0)

    workbook = load_workbook(
        file_or_path,
        data_only=True,
        read_only=False,
    )

    try:
        sheet = (
            workbook[sheet_name]
            if sheet_name in workbook.sheetnames
            else workbook.active
        )

        header_row = _find_header_row(sheet)
        columns = _build_column_map(sheet, header_row)
        max_row, _ = _sheet_size(sheet)

        rows = []

        for row_number in range(header_row + 1, max_row + 1):
            raw = {
                "code": _cell(sheet, row_number, columns, "CODIGO_PRODUCTO"),
                "description": _cell(sheet, row_number, columns, "DESCRIPCION"),
                "category": _cell(sheet, row_number, columns, "CATEGORIA"),
                "unit": _cell(sheet, row_number, columns, "UNIDAD_MEDIDA"),
                "quantity": _cell(sheet, row_number, columns, "CANTIDAD"),
                "minimum_stock": _cell(sheet, row_number, columns, "STOCK_MINIMO"),
                "notes": _cell(sheet, row_number, columns, "OBSERVACIONES"),
            }

            # Ignorar filas completamente vacías.
            if not any(value not in (None, "") for value in raw.values()):
                continue

            errors = []
            warnings = []
            review_required = False

            code = str(raw["code"]).strip().upper() if raw["code"] not in (None, "") else ""
            description = str(raw["description"]).strip() if raw["description"] not in (None, "") else ""
            category_text = str(raw["category"]).strip() if raw["category"] not in (None, "") else ""
            unit_text = str(raw["unit"]).strip() if raw["unit"] not in (None, "") else ""
            notes = str(raw["notes"]).strip() if raw["notes"] not in (None, "") else ""

            if not description:
                errors.append("DESCRIPCION es obligatoria.")
            if not category_text:
                errors.append("CATEGORIA es obligatoria.")
            if not unit_text:
                errors.append("UNIDAD_MEDIDA es obligatoria.")

            quantity = None
            minimum_stock = None
            try:
                quantity = parse_decimal(raw["quantity"], field_name="CANTIDAD", allow_zero=True)
            except ValidationError as exc:
                errors.extend(exc.messages)

            if raw["minimum_stock"] in (None, ""):
                minimum_stock = None
            else:
                try:
                    minimum_stock = parse_decimal(
                        raw["minimum_stock"],
                        field_name="STOCK_MINIMO",
                        allow_zero=True,
                    )
                except ValidationError as exc:
                    errors.extend(exc.messages)

            category = resolve_category(section=section, raw_name=category_text) if category_text else None
            unit = resolve_unit(unit_text) if unit_text else None

            if category_text and category is None:
                warnings.append(f"La categoría '{category_text}' no existe en {section.name}.")
                review_required = True

            if unit_text and unit is None:
                warnings.append(f"La unidad '{unit_text}' no existe ni coincide con un alias conocido.")
                review_required = True

            matched_product = None
            possible_products = []
            match_method = None

            if code:
                matched_product = get_product_by_code(code)
                match_method = "CODE"

                if matched_product is None:
                    errors.append(
                        f"El código '{code}' no existe. Si es un producto nuevo, deje CODIGO_PRODUCTO vacío."
                    )
                else:
                    if matched_product.section_id != section.id:
                        errors.append(
                            f"El código '{code}' pertenece a la sección "
                            f"'{matched_product.section.name}', no a '{section.name}'."
                        )
                    if description and matched_product.normalized_description != normalize_text(description):
                        warnings.append(
                            f"El código existe, pero la descripción del sistema es "
                            f"'{matched_product.description}'."
                        )
                        review_required = True
                    if unit and matched_product.unit_id != unit.id:
                        warnings.append(
                            f"El código existe, pero su unidad en el sistema es "
                            f"'{matched_product.unit.abbreviation}', no '{unit_text}'."
                        )
                        review_required = True
                    if category and matched_product.category_id != category.id:
                        current = matched_product.category.name if matched_product.category else "SIN CATEGORÍA"
                        warnings.append(
                            f"El código existe, pero su categoría actual es '{current}'. "
                            "La categoría no cambia automáticamente durante una carga."
                        )
                        review_required = True
            elif description and unit:
                matches = list(find_matching_products(
                    section=section,
                    description=description,
                    unit=unit,
                )[:20])

                if len(matches) == 1:
                    matched_product = matches[0]
                    match_method = "DESCRIPTION_UNIT"
                    if category and matched_product.category_id != category.id:
                        current = matched_product.category.name if matched_product.category else "SIN CATEGORÍA"
                        warnings.append(
                            f"Se encontró '{matched_product.code}', pero está en la categoría '{current}'."
                        )
                        review_required = True
                elif len(matches) > 1:
                    possible_products = [
                        {
                            "id": p.id,
                            "code": p.code,
                            "description": p.description,
                            "unit": p.unit.abbreviation,
                            "category": p.category.name if p.category else None,
                        }
                        for p in matches
                    ]
                    warnings.append("Hay varios productos compatibles. Debe seleccionar cuál usar.")
                    review_required = True

            if errors:
                status = "ERROR"
            elif review_required:
                status = "REVIEW"
            elif matched_product:
                status = "READY_EXISTING"
            else:
                status = "READY_NEW"

            rows.append({
                "row_number": row_number,
                "status": status,
                "code": code,
                "description": description,
                "category_raw": category_text,
                "category_id": category.id if category else None,
                "unit_raw": unit_text,
                "unit_id": unit.id if unit else None,
                "quantity": _decimal_to_string(quantity),
                "minimum_stock": _decimal_to_string(minimum_stock),
                "notes": notes,
                "matched_product_id": matched_product.id if matched_product else None,
                "matched_product_code": matched_product.code if matched_product else None,
                "match_method": match_method,
                "possible_products": possible_products,
                "warnings": warnings,
                "errors": errors,
            })

        summary = {
            "total_rows": len(rows),
            "ready_existing": sum(r["status"] == "READY_EXISTING" for r in rows),
            "ready_new": sum(r["status"] == "READY_NEW" for r in rows),
            "review": sum(r["status"] == "REVIEW" for r in rows),
            "errors": sum(r["status"] == "ERROR" for r in rows),
        }

        return {
            "sheet": sheet.title,
            "header_row": header_row,
            "section_id": section.id,
            "section_name": section.name,
            "summary": summary,
            "rows": rows,
            "can_import_without_review": summary["review"] == 0 and summary["errors"] == 0,
        }

    finally:
        workbook.close()